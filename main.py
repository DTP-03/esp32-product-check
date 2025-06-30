from flask import Flask, request, render_template, jsonify
from flask import send_file
from datetime import datetime
import pytz
import os
import cv2
import numpy as np
import pandas as pd
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage  # pip install pillow

EXCEL_PATH = os.path.join(os.getcwd(), "product_log.xlsx")

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'static'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

mode = "auto"  # auto or manual
latest_result = {"status": "WAITING", "timestamp": "", "image": ""}
last_returned_result = {"status": "", "timestamp": "", "image": ""}
bangtai_status = "START"
total_count = 0
OK_count = 0
# Múi giờ Việt Nam
VN_TZ = pytz.timezone("Asia/Ho_Chi_Minh")

def detect_defect(img_path):
    img = cv2.imread(img_path)
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    # Ngưỡng màu đỏ
    lower_red1 = np.array([0, 100, 100])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([160, 100, 100])
    upper_red2 = np.array([179, 255, 255])

    # Tạo mask màu đỏ
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask_red = cv2.bitwise_or(mask1, mask2)

    # Làm mượt mask để tránh nhiễu
    blurred = cv2.GaussianBlur(mask_red, (9, 9), 2)

    # Phát hiện hình tròn trong vùng đỏ
    circles = cv2.HoughCircles(
        blurred, 
        cv2.HOUGH_GRADIENT, 
        dp=1.2, 
        minDist=30,
        param1=100,
        param2=30,  # càng thấp thì càng dễ phát hiện (nhiễu)
        minRadius=10, 
        maxRadius=150
    )

    if circles is not None:
        return "OK"  # có hình tròn màu đỏ
    else:
        return "ERROR"

@app.route('/')
def index():
    return render_template('index.html', time=datetime.now(VN_TZ).timestamp())

@app.route('/upload', methods=['POST'])
def upload_image():
    global latest_result, total_count, OK_count

    now = datetime.now(VN_TZ).strftime("%Y%m%d-%H%M%S")
    filename = f"{now}.jpg"
    filepath = os.path.join(UPLOAD_FOLDER, filename)

    with open(filepath, 'wb') as f:
        f.write(request.data)
        f.flush()
        os.fsync(f.fileno())

    if mode == "auto":
        result = detect_defect(filepath)
    else:
        result = latest_result.get("status", "OK")
        
      # cộng số lượng các sản phẩm  
    total_count +=1;
    if result == "OK":
        OK_count +=1;
    
    latest_result = {"status": result, "timestamp": now, "image": filename}

    # === Ghi vào Excel với ảnh ===
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Timestamp", "Status", "Image"])  # Header
        wb.save(EXCEL_PATH)
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    
    row = ws.max_row + 1
    # Ghi dữ liệu
    ws.cell(row=row, column=1, value=now)
    ws.cell(row=row, column=2, value=result)
    ws.cell(row=row, column=3, value=filename)  # GHI TÊN FILE vào ô C để dùng sau
    
    # Chèn ảnh vào vị trí C (vẫn là ô C nhưng ảnh đè lên)
    img = XLImage(filepath)
    img.width = 100
    img.height = 100
    img.anchor = f"D{row}"
    ws.add_image(img)
    
    # Điều chỉnh kích thước
    ws.row_dimensions[row].height = 80
    ws.column_dimensions["D"].width = 18

    wb.save(EXCEL_PATH)
    print(f"Đã ghi sản phẩm {result} với ảnh {filename} vào Excel.")

    return jsonify({"result": result})



@app.route('/manual-result', methods=['POST'])
def manual_result():
    global latest_result
    result = request.json.get("result")
    if result in ["OK", "ERROR"]:
        latest_result["status"] = result
        return jsonify({"status": result})
    return jsonify({"error": "Invalid result"}), 400

@app.route('/status')
def get_status():
    global last_returned_result
    if latest_result["status"] != "WAITING":
        last_returned_result = latest_result
    return jsonify(last_returned_result)

@app.route('/set-bangtai', methods=['POST'])
def set_bangtai():
    global bangtai_status
    data = request.get_json()
    if data and data.get("Bangtai") in ["START", "STOP"]:
        bangtai_status = data["Bangtai"]
        return jsonify({"Bangtai": bangtai_status})
    return jsonify({"error": "Invalid command"}), 400

@app.route('/bangtai')
def bangtai():
    return jsonify({"Bangtai": bangtai_status})


@app.route("/images", methods=["GET"])
def get_images():
    try:
        if not os.path.exists(EXCEL_PATH):
            return jsonify([])

        df = pd.read_excel(EXCEL_PATH)
        images = []
        for _, row in df.iterrows():
            images.append({
                "url": f"/static/{row['Image']}",
                "tags": [row['Status']],
                "created_at": datetime.strptime(str(row['Timestamp']), "%Y%m%d-%H%M%S").isoformat()
            })
        return jsonify(images)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/counts')
def get_counts():
    return jsonify({
        "total": total_count,
        "ok": OK_count,
    })
    
@app.route('/export-excel')
def export_excel():
    if os.path.exists(EXCEL_PATH):
        return send_file(EXCEL_PATH, as_attachment=True)
    return jsonify({"error": "Excel file not found"}), 404

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)
